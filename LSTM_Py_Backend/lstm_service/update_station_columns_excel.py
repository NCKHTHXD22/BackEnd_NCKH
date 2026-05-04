"""
update_station_columns_excel.py
================================
Cập nhật phần dữ liệu từng trạm riêng lẻ (cols 122+) trong tất cả 816 file Excel
Data_Tung_Ho_Ma_Tran_Rong dựa trên bộ trạm MỚI trong rain_stations.py.

Thay đổi:
  - Xóa toàn bộ station columns cũ (cols 122+)
  - Ghi lại đầy đủ tất cả trạm trong RESERVOIR_TO_STATIONS[rid]
  - Trạm có data → giá trị thực từ VNDMS
  - Trạm không có data tháng đó → 0.0 (không bỏ trống)

Cấu trúc mỗi station block (24 cột):
  Row 1: "Trạm: {tên hiển thị}"
  Row 2: 00h | 01h | ... | 23h
  Row 3: (metadata — để trống)
  Row 4+: giá trị mưa theo giờ

Chạy từ thư mục lstm_service:
    python -X utf8 update_station_columns_excel.py
"""
import os
import sys
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from config.rain_stations import RAIN_STATIONS, RESERVOIR_TO_STATIONS
from config.reservoirs import RESERVOIRS
from data.vndms_rain_loader import _parse_vndms_excel, _find_station_rows

RAIN_DATA_ROOT = os.path.join(SCRIPT_DIR, "Rain Data")
EXCEL_ROOT     = os.path.join(SCRIPT_DIR, "LSTM_Project", "Data_Tung_Ho_Ma_Tran_Rong")

MONTHS = [
    (yr, mn)
    for yr in range(2022, 2027)
    for mn in range(1, 13)
    if (yr, mn) <= (2026, 3)
]

# Cột bắt đầu của phần station data (1-indexed openpyxl)
# Col 1=Date, 2-25=Z, 26-49=Inflow, 50-73=Qout, 74-97=Rain IDW, 98-121=P50
STATION_COL_START = 122
HEADER_ROWS       = 3      # rows 1,2,3 là header
DATA_ROW_START    = 4      # row 4 là dòng dữ liệu đầu tiên
HOURS             = 24


def _get_station_display_name(sk: str) -> str:
    """Tên hiển thị ngắn gọn cho header Excel."""
    info = RAIN_STATIONS.get(sk, {})
    desc = info.get("description", sk)
    # Lấy phần đầu trước dấu " —"
    return desc.split(" —")[0].strip() if " —" in desc else sk


def _get_vndms_display_name(vndms_df: pd.DataFrame, sk: str) -> str:
    """Tên trạm thực tế trong VNDMS (ưu tiên tên đầy đủ từ file)."""
    rows = _find_station_rows(vndms_df, sk)
    if rows.empty:
        return _get_station_display_name(sk)
    names = rows["station"].dropna().unique()
    if len(names) == 0:
        return _get_station_display_name(sk)
    return str(max(names, key=len))


def compute_station_daily(vndms_df: pd.DataFrame, sk: str,
                          year: int, month: int) -> dict:
    """
    Trả về {date: array(24,)} — mưa theo giờ cho trạm sk trong tháng year/month.
    Trạm không có data → array toàn 0.0
    """
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    t_start  = pd.Timestamp(year, month, 1, 0)
    t_end    = pd.Timestamp(year, month, last_day, 23)
    time_idx = pd.date_range(t_start, t_end, freq="h")

    rows = _find_station_rows(vndms_df, sk)
    if rows.empty:
        # Không có data → trả 0 cho mọi giờ
        result = {}
        for d in range(1, last_day + 1):
            result[datetime(year, month, d).date()] = np.zeros(24, dtype=np.float32)
        return result

    ts = rows.groupby("time")["rain"].mean()
    ts = ts.reindex(time_idx).fillna(0.0)

    result = {}
    for d in range(1, last_day + 1):
        idx = (d - 1) * 24
        result[datetime(year, month, d).date()] = ts.values[idx:idx+24].astype(np.float32)
    return result


def update_station_columns(xlsx_path: str, vndms_df: pd.DataFrame,
                            rid: int, year: int, month: int) -> str:
    """
    Cập nhật phần station columns (col 122+) của 1 file Excel.
    Trả về chuỗi mô tả kết quả.
    """
    station_keys = RESERVOIR_TO_STATIONS.get(rid, [])
    if not station_keys:
        return "SKIP: no stations"

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # ── 1. Xóa toàn bộ cols 122+ ─────────────────────────────────────────
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=STATION_COL_START, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # ── 2. Lấy mapping ngày → row index ──────────────────────────────────
    date_to_row = {}
    for r in range(DATA_ROW_START, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        if isinstance(val, datetime):
            d = val.date()
        else:
            try:
                d = pd.to_datetime(str(val)).date()
            except Exception:
                continue
        date_to_row[d] = r

    stations_written = []

    # ── 3. Ghi từng station block ─────────────────────────────────────────
    for i, sk in enumerate(station_keys):
        col_start = STATION_COL_START + i * HOURS   # col đầu của block này

        # Lấy tên hiển thị
        display_name = _get_vndms_display_name(vndms_df, sk)
        available_from = RAIN_STATIONS.get(sk, {}).get("available_from", "?")

        # Lấy data cho tháng này
        daily_data = compute_station_daily(vndms_df, sk, year, month)

        # Row 1: tên trạm
        cell_r1 = ws.cell(1, col_start)
        cell_r1.value = f"Trạm: {display_name}"
        cell_r1.font  = Font(bold=True)

        # Row 2: hour labels
        for h in range(HOURS):
            ws.cell(2, col_start + h).value = f"{h:02d}h"

        # Row 3: metadata (available_from)
        ws.cell(3, col_start).value = f"({available_from}+)"

        # Row 4+: data
        for date, arr in sorted(daily_data.items()):
            row_idx = date_to_row.get(date)
            if row_idx is None:
                continue
            for h in range(HOURS):
                val = float(arr[h])
                ws.cell(row_idx, col_start + h).value = round(val, 3) if val != 0.0 else 0

        stations_written.append(sk)

    wb.save(xlsx_path)
    wb.close()
    return f"{len(stations_written)} trams [{','.join(stations_written)}]"


def main():
    print("=" * 70)
    print("UPDATE STATION COLUMNS — TẤT CẢ FILE EXCEL")
    print(f"Tổng tháng: {len(MONTHS)} | Tổng hồ: {len(RESERVOIRS)}")
    print("=" * 70)

    total_ok = total_err = total_skip = 0

    for year, month in MONTHS:
        vndms_path = os.path.join(
            RAIN_DATA_ROOT,
            f"DATA_RAIN_{year}",
            f"VNDMS_Mua_Theo_Gio_{year}_{month:02d}.xlsx"
        )
        if not os.path.exists(vndms_path):
            print(f"\n[SKIP] VNDMS không tồn tại: {vndms_path}")
            total_skip += len(RESERVOIRS)
            continue

        print(f"\n{'─'*60}")
        print(f"[{year}-{month:02d}] Load VNDMS...")
        vndms_df = _parse_vndms_excel(vndms_path)
        if vndms_df.empty:
            total_skip += len(RESERVOIRS)
            continue

        n_stations_present = vndms_df["station"].nunique()
        print(f"  Trạm có dữ liệu: {n_stations_present}")

        for rid, res_info in sorted(RESERVOIRS.items()):
            ho_name   = res_info["name"].replace(" ", "_")
            xlsx_name = f"{ho_name}_{year}_{month:02d}.xlsx"
            xlsx_path = os.path.join(EXCEL_ROOT, ho_name, xlsx_name)

            if not os.path.exists(xlsx_path):
                total_skip += 1
                continue

            try:
                result = update_station_columns(xlsx_path, vndms_df, rid, year, month)
                print(f"  rid={rid:2d} {ho_name:25s}: {result}")
                total_ok += 1
            except Exception as e:
                print(f"  [ERROR] rid={rid} {ho_name}: {e}")
                total_err += 1

    print("\n" + "=" * 70)
    print(f"HOÀN THÀNH: OK={total_ok} | SKIP={total_skip} | ERROR={total_err}")
    print("=" * 70)


if __name__ == "__main__":
    main()
