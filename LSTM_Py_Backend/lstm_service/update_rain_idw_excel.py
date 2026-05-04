"""
update_rain_idw_excel.py
========================
Cập nhật cột mưa IDW (cols 73-96) trong tất cả 816 file Excel
Data_Tung_Ho_Ma_Tran_Rong dựa trên bộ trạm mưa MỚI trong rain_stations.py.

Logic chính:
  - Mỗi tháng: load 1 file VNDMS_Mua_Theo_Gio_YYYY_MM.xlsx
  - Cho từng hồ: tính IDW từ các trạm CÓ DỮ LIỆU trong tháng đó (renormalize tự động)
  - Cập nhật cột rain IDW trong file Excel tương ứng (openpyxl, in-place)

Chạy từ thư mục lstm_service:
    python -X utf8 update_rain_idw_excel.py
"""
import os
import sys
import glob
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime

# ── Path setup ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from config.rain_stations import RAIN_STATIONS, RESERVOIR_TO_STATIONS
from config.reservoirs import RESERVOIRS
from data.idw_calculator import compute_idw_weights
from data.vndms_rain_loader import _parse_vndms_excel, _find_station_rows

# ── Cấu hình đường dẫn ──────────────────────────────────────────────────────
RAIN_DATA_ROOT = os.path.join(SCRIPT_DIR, "Rain Data")
EXCEL_ROOT     = os.path.join(SCRIPT_DIR, "LSTM_Project", "Data_Tung_Ho_Ma_Tran_Rong")

# Tháng cần xử lý: 2022-01 → 2026-03
MONTHS = [
    (yr, mn)
    for yr in range(2022, 2027)
    for mn in range(1, 13)
    if (yr, mn) <= (2026, 3)
]

# Hàng header trong file Excel (openpyxl 1-indexed)
HEADER_ROWS = 3        # row 1, 2, 3 là header
DATA_ROW_START = 4     # row 4 là ngày đầu tiên

# Cột mưa IDW trong schema MỚI (0-indexed: 73-96 → openpyxl 1-indexed: 74-97)
RAIN_COL_START = 74    # openpyxl column index của giờ 00h
RAIN_COL_END   = 97    # openpyxl column index của giờ 23h (inclusive)
HOURS          = 24


# ────────────────────────────────────────────────────────────────────────────
#  Hàm detect schema
# ────────────────────────────────────────────────────────────────────────────
def _detect_rain_col_start(ws) -> int:
    """
    Trả về cột bắt đầu mưa IDW (1-indexed openpyxl).
    Schema mới: col 74 (0-indexed: 73)
    Schema cũ : col 26 (0-indexed: 25)
    """
    row1_vals = [str(ws.cell(row=1, column=c).value or "").lower() for c in range(1, 130)]
    has_z = any("m\u1ef1c n\u01b0\u1edbc" in v or "muc nuoc" in v for v in row1_vals)
    return 74 if has_z else 26


# ────────────────────────────────────────────────────────────────────────────
#  Tính IDW mưa cho 1 hồ trong 1 tháng
# ────────────────────────────────────────────────────────────────────────────
def compute_monthly_idw(vndms_df: pd.DataFrame, rid: int,
                        year: int, month: int) -> dict:
    """
    Tính chuỗi mưa IDW (hourly) cho hồ rid trong tháng year/month.

    Returns:
        dict { date: np.ndarray(24,) } — mưa theo 24 giờ cho mỗi ngày
        (date là Python date object)
    """
    # Trọng số IDW cho tất cả trạm (dựa trên khoảng cách đến tâm hồ)
    all_weights = compute_idw_weights(rid)
    if not all_weights:
        return {}

    # Lọc timestamp thuộc tháng này
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    t_start  = pd.Timestamp(year, month, 1, 0)
    t_end    = pd.Timestamp(year, month, last_day, 23)
    time_idx = pd.date_range(t_start, t_end, freq="h")

    # Lấy timeseries cho từng trạm (có trong tháng này)
    station_series = {}
    station_weights_used = {}
    for sk, w in all_weights.items():
        rows = _find_station_rows(vndms_df, sk)
        if rows.empty:
            continue
        ts = rows.groupby("time")["rain"].mean()
        ts = ts.reindex(time_idx)      # align to hourly index, NaN nếu không có
        if ts.isna().all():
            continue                   # trạm hoàn toàn không có dữ liệu tháng này
        station_series[sk]         = ts.values  # np.array (n_hours,)
        station_weights_used[sk]   = w

    if not station_series:
        return {}

    # Build matrix (n_stations, n_hours)
    sk_list   = list(station_series.keys())
    w_arr     = np.array([station_weights_used[sk] for sk in sk_list], dtype=np.float64)
    rain_mat  = np.stack([station_series[sk] for sk in sk_list], axis=0)  # (n_sta, n_hrs)

    # IDW với renormalize theo từng giờ (chỉ dùng trạm có dữ liệu tại giờ đó)
    not_nan   = ~np.isnan(rain_mat)                            # (n_sta, n_hrs)
    w_avail   = np.where(not_nan, w_arr[:, None], 0.0)        # (n_sta, n_hrs)
    w_sum     = w_avail.sum(axis=0)                            # (n_hrs,)
    rain_sum  = np.where(not_nan, rain_mat * w_arr[:, None], 0.0).sum(axis=0)  # (n_hrs,)
    rain_idw  = np.where(w_sum > 0, rain_sum / w_sum, 0.0)    # (n_hrs,)

    # Tổ chức kết quả theo ngày → {date: array(24,)}
    result = {}
    for d in range(1, last_day + 1):
        day_start_idx = (d - 1) * 24
        arr = rain_idw[day_start_idx : day_start_idx + 24]
        result[datetime(year, month, d).date()] = arr.astype(np.float32)

    return result


# ────────────────────────────────────────────────────────────────────────────
#  Cập nhật file Excel
# ────────────────────────────────────────────────────────────────────────────
def update_excel_rain(xlsx_path: str, daily_rain: dict, year: int, month: int) -> int:
    """
    Cập nhật cột mưa IDW trong file Excel xlsx_path.

    Args:
        xlsx_path : đường dẫn tuyệt đối đến file Excel
        daily_rain: {date: array(24,)} — mưa IDW theo ngày/giờ
        year, month: để xác định rows

    Returns:
        Số ngày đã cập nhật
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    rain_col_start = _detect_rain_col_start(ws)

    updated = 0
    # Duyệt từng data row
    for row in ws.iter_rows(min_row=DATA_ROW_START, values_only=False):
        date_cell = row[0]
        date_val  = date_cell.value

        # Parse date cell
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            d = date_val.date()
        elif hasattr(date_val, 'date'):
            d = date_val.date()
        else:
            try:
                d = pd.to_datetime(str(date_val)).date()
            except Exception:
                continue

        if d.year != year or d.month != month:
            continue

        arr = daily_rain.get(d)
        if arr is None:
            continue

        # Ghi 24 giá trị mưa vào các cột rain
        for h in range(24):
            ws.cell(row=date_cell.row, column=rain_col_start + h).value = round(float(arr[h]), 3)
        updated += 1

    wb.save(xlsx_path)
    wb.close()
    return updated


# ────────────────────────────────────────────────────────────────────────────
#  Main
# ────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("UPDATE MƯA IDW — TẤT CẢ FILE EXCEL Data_Tung_Ho_Ma_Tran_Rong")
    print(f"Tổng tháng cần xử lý : {len(MONTHS)}")
    print(f"Tổng hồ               : {len(RESERVOIRS)}")
    print(f"Tổng file Excel       : {len(MONTHS) * len(RESERVOIRS)}")
    print("=" * 70)

    total_ok  = 0
    total_err = 0
    total_skip = 0

    for year, month in MONTHS:
        # ── Load file VNDMS tháng này ────────────────────────────────────
        vndms_path = os.path.join(
            RAIN_DATA_ROOT,
            f"DATA_RAIN_{year}",
            f"VNDMS_Mua_Theo_Gio_{year}_{month:02d}.xlsx"
        )
        if not os.path.exists(vndms_path):
            print(f"\n[SKIP] VNDMS không tồn tại: {vndms_path}")
            total_skip += len(RESERVOIRS)
            continue

        print(f"\n{'─'*60}")
        print(f"[{year}-{month:02d}] Load VNDMS: {os.path.basename(vndms_path)}")
        vndms_df = _parse_vndms_excel(vndms_path)
        if vndms_df.empty:
            print(f"  [WARN] VNDMS rỗng, bỏ qua tháng này")
            total_skip += len(RESERVOIRS)
            continue

        # Trạm có mặt trong tháng này
        present_stations = set(vndms_df["station"].unique())
        print(f"  Trạm có dữ liệu: {len(present_stations)}")

        # ── Xử lý từng hồ ───────────────────────────────────────────────
        for rid, res_info in sorted(RESERVOIRS.items()):
            ho_name   = res_info["name"].replace(" ", "_")
            xlsx_name = f"{ho_name}_{year}_{month:02d}.xlsx"
            xlsx_path = os.path.join(EXCEL_ROOT, ho_name, xlsx_name)

            if not os.path.exists(xlsx_path):
                print(f"  [SKIP] {xlsx_name} không tồn tại")
                total_skip += 1
                continue

            # Tính IDW cho hồ này trong tháng này
            all_w = compute_idw_weights(rid)
            stations_used = [sk for sk in all_w if not _find_station_rows(vndms_df, sk).empty]

            if not stations_used:
                print(f"  [WARN] rid={rid} ({ho_name}): không có trạm nào có dữ liệu → skip")
                total_skip += 1
                continue

            try:
                daily_rain = compute_monthly_idw(vndms_df, rid, year, month)
                if not daily_rain:
                    print(f"  [WARN] rid={rid}: IDW trả rỗng → skip")
                    total_skip += 1
                    continue

                n_days = update_excel_rain(xlsx_path, daily_rain, year, month)
                status = f"{n_days}d OK [{','.join(stations_used)}]"
                print(f"  rid={rid:2d} {ho_name:25s}: {status}")
                total_ok += 1

            except Exception as e:
                print(f"  [ERROR] rid={rid} {ho_name}: {e}")
                total_err += 1

    print("\n" + "=" * 70)
    print(f"HOÀN THÀNH: OK={total_ok} | SKIP={total_skip} | ERROR={total_err}")
    print("=" * 70)


if __name__ == "__main__":
    main()
