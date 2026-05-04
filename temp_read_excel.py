import openpyxl
import sys

# Read a sample Excel file
file_path = r"LSTM_Py_Backend\lstm_service\LSTM_Project\Data_Tung_Ho_Ma_Tran_Rong\HO_SONG_TRANH_2\HO_SONG_TRANH_2_2022_01.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    
    print("=== File Structure ===")
    print(f"Active sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Read row 0 (features)
    print("\nRow 0 (Features):")
    for col_idx in range(1, min(ws.max_column + 1, 130)):
        cell_val = ws.cell(row=1, column=col_idx).value
        if col_idx <= 5 or (col_idx >= 70 and col_idx <= 100):
            print(f"  Col {col_idx}: {cell_val}")
    
    # Read row 1 (Hour labels)
    print("\nRow 1 (Hour labels - first 30):")
    for col_idx in range(1, 31):
        cell_val = ws.cell(row=2, column=col_idx).value
        print(f"  Col {col_idx}: {cell_val}")
    
    # Read first data row
    print("\nRow 2 (First data row - first 30):")
    for col_idx in range(1, 31):
        cell_val = ws.cell(row=3, column=col_idx).value
        print(f"  Col {col_idx}: {cell_val}")
        
except Exception as e:
    print(f"Error: {e}", file=sys.stderr)
    import traceback
    traceback.print_exc()

